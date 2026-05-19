import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import re

st.set_page_config(page_title="Generador IQVIA", page_icon="💊", layout="wide")
st.title("💊 Generador de Mercado IQVIA")
st.caption("Sube la sábana mensual · selecciona criterios · descarga el Excel dinámico con filtros de país y forma de administración.")

# ── ESTILOS EXCEL ─────────────────────────────────────────────────────────
DARK_BLUE  = PatternFill('solid', start_color='17375E')
MED_BLUE   = PatternFill('solid', start_color='1F4E79')
ALT_FILL   = PatternFill('solid', start_color='DEEAF1')
WHITE_FILL = PatternFill('solid', start_color='FFFFFF')
TOTAL_FILL = PatternFill('solid', start_color='D9E1F2')
GRAY_FILL  = PatternFill('solid', start_color='F2F2F2')
YELLOW     = PatternFill('solid', start_color='FFFF99')

WHITE_BOLD  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
DARK_BOLD   = Font(name='Arial', bold=True, color='000000', size=10)
DARK_NORM   = Font(name='Arial', size=9,  color='000000')
DARK_NORM10 = Font(name='Arial', size=10, color='000000')

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right',  vertical='center')
thin   = Side(style='thin', color='B8CCE4')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, r, c, v, fill=DARK_BLUE, font=WHITE_BOLD, align=CENTER):
    cell = ws.cell(r, c, v)
    cell.fill = fill; cell.font = font; cell.alignment = align; cell.border = BORDER
    return cell

def dat(ws, r, c, v, font=DARK_NORM, align=LEFT, fill=None, fmt=None):
    cell = ws.cell(r, c, v)
    cell.font = font; cell.alignment = align; cell.border = BORDER
    if fill: cell.fill = fill
    if fmt:  cell.number_format = fmt
    return cell

# ── PARSEO DE MAT ────────────────────────────────────────────────────────
MONTH_MAP = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
             'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}

def parse_col_date(col_name):
    m = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{4})', str(col_name))
    if m:
        return int(m.group(2)), MONTH_MAP[m.group(1)], m.group(1)
    return None

def compute_mat_periods(unit_cols, usd_cols):
    """
    MAT siempre cierra en el mismo mes que el último disponible en la data.
    Ej: si el último mes es Feb 2026:
        MAT Feb 2022 = Mar 2021 → Feb 2022
        MAT Feb 2023 = Mar 2022 → Feb 2023
        MAT Feb 2024 = Mar 2023 → Feb 2024
        MAT Feb 2025 = Mar 2024 → Feb 2025
        MAT Feb 2026 = Mar 2025 → Feb 2026
    """
    dated = []
    for uc, dc in zip(unit_cols, usd_cols):
        parsed = parse_col_date(uc)
        if parsed:
            yr, mo, mo_name = parsed
            dated.append(((yr, mo), mo_name, uc, dc))
    dated.sort(key=lambda x: x[0])

    if len(dated) < 12:
        return []

    # Detectar el último mes disponible
    last_mo = dated[-1][0][1]  # número de mes del último disponible

    # Buscar todos los índices donde el mes == last_mo
    mat_indices = [i for i, ((yr, mo), _, _, _) in enumerate(dated) if mo == last_mo]

    periods = []
    for end_i in mat_indices:
        if end_i < 11:
            continue
        window = dated[end_i - 11: end_i + 1]
        if len(window) != 12:
            continue
        (yr, mo), mo_name, _, _ = dated[end_i]
        label   = f"MAT {mo_name} {yr}"
        ucols12 = [w[2] for w in window]
        dcols12 = [w[3] for w in window]
        periods.append((label, ucols12, dcols12))

    return periods

# ── CARGA DE DATOS ────────────────────────────────────────────────────────
@st.cache_data(show_spinner="Leyendo sábana IQVIA... puede tardar unos segundos.")
def cargar_sabana(file_bytes):
    df = pd.read_excel(BytesIO(file_bytes), header=None)
    headers_raw = list(df.iloc[0])
    df = df.iloc[1:].reset_index(drop=True)
    df.columns = headers_raw
    str_cols = ['Country Desc','Atc I','Atc IV','Molecule Desc','Prod Desc',
                'Pack Desc','Manu Desc','Pack Mark Desc','Pack Gene Desc','App1 Desc']
    for col in str_cols:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    unit_cols = [c for c in df.columns if 'Sales Units Qty'          in str(c)]
    usd_cols  = [c for c in df.columns if 'Sales Value List Usd Amt'  in str(c)]
    for col in unit_cols + usd_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    return df, unit_cols, usd_cols

def fmt_launch(val):
    try:
        if isinstance(val, datetime): return val.strftime('%d/%m/%Y')
        if isinstance(val, (int, float)) and not np.isnan(float(val)):
            return (datetime(1899,12,30)+pd.Timedelta(days=int(val))).strftime('%d/%m/%Y')
        if isinstance(val, str) and val not in ('nan','NaT','','None'): return val
    except: pass
    return ''

# ── GENERADOR DE EXCEL ────────────────────────────────────────────────────
def generar_excel(df_atc, df_molecula, mat_periods):
    wb = Workbook()

    yr_short, yr_full = [], []
    for lbl, _, _ in mat_periods:
        parts = lbl.replace('MAT ','').split()
        yr_full.append(f"{parts[0]} {parts[1]}")
        yr_short.append(parts[1])

    last_yr = yr_short[-1]
    prev_yr = yr_short[-2] if len(yr_short) >= 2 else None

    # Calcular MAT
    for df in [df_atc, df_molecula]:
        for lbl, ucols, dcols in mat_periods:
            yr = lbl.replace('MAT ','').split()[-1]
            df[f'UNI {yr}'] = df[ucols].sum(axis=1)
            df[f'USD {yr}'] = df[dcols].sum(axis=1)

    uni_mat   = [f'UNI {y}' for y in yr_short]
    usd_mat   = [f'USD {y}' for y in yr_short]
    atc_label = df_atc['Atc IV'].mode()[0] if len(df_atc) else ''

    paises   = sorted(df_atc['Country Desc'].dropna().unique().tolist())
    formas   = sorted(df_atc['App1 Desc'].dropna().unique().tolist())
    n_paises = len(paises) + 1
    n_formas = len(formas) + 1

    # ── HOJA _LISTAS (oculta) ─────────────────────────────────────────────
    ws_listas = wb.create_sheet('_LISTAS')
    ws_listas.sheet_state = 'hidden'
    ws_listas['A1'] = '(Todos)'
    for i, p in enumerate(paises, 2): ws_listas.cell(i, 1, p)
    ws_listas['B1'] = '(Todas)'
    for i, f in enumerate(formas, 2): ws_listas.cell(i, 2, f)

    # ── HOJA DATA (ATC IV completo) ───────────────────────────────────────
    ws_data = wb.active
    ws_data.title = 'DATA'

    info_cols = ['Country Desc','Atc I','Atc IV','Molecule Desc','Prod Desc','Pack Desc',
                 'Manu Desc','Pack Mark Desc','Pack Gene Desc','App1 Desc','Pack Launch Dt']

    # Posiciones de columnas clave en DATA (1-based)
    COL_PAIS  = 1   # A
    COL_MOL   = 4   # D
    COL_PROD  = 5   # E
    COL_PACK  = 6   # F
    COL_FORMA = 10  # J

    h_data = info_cols + \
             [f'Sales Units Qty\nMAT {yf}'         for yf in yr_full] + \
             [f'Sales Value List Usd Amt\nMAT {yf}' for yf in yr_full]

    for ci, h in enumerate(h_data, 1):
        hdr(ws_data, 1, ci, h)
    ws_data.row_dimensions[1].height = 42

    for ri, (_, row) in enumerate(df_atc.iterrows(), 2):
        fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        for ci, col in enumerate(info_cols, 1):
            v = row.get(col,'')
            if col == 'Pack Launch Dt': v = fmt_launch(v)
            dat(ws_data, ri, ci, v if str(v) not in ('nan','None') else '', fill=fill)
        for ci, col in enumerate(uni_mat, len(info_cols)+1):
            dat(ws_data, ri, ci, round(float(row.get(col,0)),0), align=RIGHT, fill=fill, fmt='#,##0')
        for ci, col in enumerate(usd_mat, len(info_cols)+len(uni_mat)+1):
            dat(ws_data, ri, ci, round(float(row.get(col,0)),3), align=RIGHT, fill=fill, fmt='#,##0.000')

    n_data_rows   = len(df_atc)
    data_last_row = n_data_rows + 1

    ws_data.freeze_panes = 'A2'
    ws_data.auto_filter.ref = f'A1:{get_column_letter(len(h_data))}1'
    for i, w in enumerate([14,22,30,42,20,28,15,11,9,22,13]+[13]*len(mat_periods)*2, 1):
        ws_data.column_dimensions[get_column_letter(i)].width = w

    uni_start_ci = len(info_cols)+1
    usd_start_ci = len(info_cols)+len(uni_mat)+1
    uni_letters  = [get_column_letter(uni_start_ci+i) for i in range(len(yr_short))]
    usd_letters  = [get_column_letter(usd_start_ci+i) for i in range(len(yr_short))]

    L_PAIS  = get_column_letter(COL_PAIS)
    L_MOL   = get_column_letter(COL_MOL)
    L_PROD  = get_column_letter(COL_PROD)
    L_PACK  = get_column_letter(COL_PACK)
    L_FORMA = get_column_letter(COL_FORMA)

    def add_dropdown(ws, cell_ref, lista_col, n_items):
        dv = DataValidation(type='list',
                            formula1=f'_LISTAS!${lista_col}$1:${lista_col}${n_items}',
                            showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(cell_ref)

    def filter_header(ws):
        ws['A1'] = 'PAIS';        ws['A1'].font=DARK_BOLD; ws['A1'].fill=GRAY_FILL; ws['A1'].alignment=LEFT
        ws['B1'] = '(Todos)';     ws['B1'].font=DARK_BOLD; ws['B1'].fill=YELLOW;    ws['B1'].alignment=LEFT; ws['B1'].border=BORDER
        ws['A2'] = 'FORMA ADMON'; ws['A2'].font=DARK_BOLD; ws['A2'].fill=GRAY_FILL; ws['A2'].alignment=LEFT
        ws['B2'] = '(Todas)';     ws['B2'].font=DARK_BOLD; ws['B2'].fill=YELLOW;    ws['B2'].alignment=LEFT; ws['B2'].border=BORDER
        ws['A3'] = '⚠ Cambia las celdas amarillas para filtrar. Los totales se actualizan automáticamente.'
        ws['A3'].font = Font(name='Arial', size=9, color='7F6000', italic=True)
        add_dropdown(ws, 'B1', 'A', n_paises)
        add_dropdown(ws, 'B2', 'B', n_formas)

    def sumifs_ct(ws_name, val_col_letter, mol_cell_ref):
        """SUMIFS para Mercado CT — molécula referenciada por celda"""
        pais_ref  = f"'{ws_name}'!$B$1"
        forma_ref = f"'{ws_name}'!$B$2"
        data_val  = f"DATA!${val_col_letter}$2:DATA!${val_col_letter}${data_last_row}"
        return (
            f"=SUMIFS({data_val},"
            f"DATA!${L_MOL}$2:DATA!${L_MOL}${data_last_row},{mol_cell_ref},"
            f"DATA!${L_PAIS}$2:DATA!${L_PAIS}${data_last_row},IF({pais_ref}=\"(Todos)\",\"*\",{pais_ref}),"
            f"DATA!${L_FORMA}$2:DATA!${L_FORMA}${data_last_row},IF({forma_ref}=\"(Todas)\",\"*\",{forma_ref}))"
        )

    def sumifs_mol(ws_name, val_col_letter, mol_ref, prod_ref, pack_ref):
        """SUMIFS para Mercado Molécula — mol, prod y pack referenciados por celda"""
        pais_ref  = f"'{ws_name}'!$B$1"
        forma_ref = f"'{ws_name}'!$B$2"
        data_val  = f"DATA!${val_col_letter}$2:DATA!${val_col_letter}${data_last_row}"
        return (
            f"=SUMIFS({data_val},"
            f"DATA!${L_MOL}$2:DATA!${L_MOL}${data_last_row},{mol_ref},"
            f"DATA!${L_PROD}$2:DATA!${L_PROD}${data_last_row},{prod_ref},"
            f"DATA!${L_PACK}$2:DATA!${L_PACK}${data_last_row},{pack_ref},"
            f"DATA!${L_PAIS}$2:DATA!${L_PAIS}${data_last_row},IF({pais_ref}=\"(Todos)\",\"*\",{pais_ref}),"
            f"DATA!${L_FORMA}$2:DATA!${L_FORMA}${data_last_row},IF({forma_ref}=\"(Todas)\",\"*\",{forma_ref}))"
        )

    # ── HOJA 02_Mercado CT ────────────────────────────────────────────────
    ws2 = wb.create_sheet('02_Mercado CT')
    filter_header(ws2)

    ct_hdrs = ['MERCADO','MOLECULA'] + \
              [f'UNI {y}' for y in yr_short] + \
              [f'US$ {y}' for y in yr_short] + ['crec%']
    for ci, h in enumerate(ct_hdrs, 1):
        hdr(ws2, 5, ci, h)
    ws2.row_dimensions[5].height = 30

    mols_atc = df_atc.groupby('Molecule Desc')[uni_mat+usd_mat].sum()
    mols_atc = mols_atc.sort_values(f'USD {last_yr}', ascending=False).reset_index()

    ri = 6
    first = True
    for _, mrow in mols_atc.iterrows():
        fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        dat(ws2, ri, 1, atc_label if first else '', font=DARK_BOLD if first else DARK_NORM, fill=fill)
        # Molécula siempre en col B — las fórmulas referencian esta celda
        dat(ws2, ri, 2, mrow['Molecule Desc'], fill=fill)
        first = False

        mol_ref = f'$B{ri}'
        for ci, ul in enumerate(uni_letters, 3):
            c = ws2.cell(ri, ci, sumifs_ct('02_Mercado CT', ul, mol_ref))
            c.font=DARK_NORM; c.alignment=RIGHT; c.fill=fill; c.border=BORDER; c.number_format='#,##0'
        for ci, dl in enumerate(usd_letters, 3+len(yr_short)):
            c = ws2.cell(ri, ci, sumifs_ct('02_Mercado CT', dl, mol_ref))
            c.font=DARK_NORM; c.alignment=RIGHT; c.fill=fill; c.border=BORDER; c.number_format='#,##0.000'
        if prev_yr:
            col_p = get_column_letter(3+len(yr_short)+yr_short.index(prev_yr))
            col_l = get_column_letter(3+len(yr_short)+yr_short.index(last_yr))
            c = ws2.cell(ri, 3+2*len(yr_short),
                f'=IF({col_p}{ri}=0,"",({col_l}{ri}-{col_p}{ri})/{col_p}{ri})')
            c.font=DARK_NORM; c.alignment=RIGHT; c.fill=fill; c.border=BORDER; c.number_format='0.00%'
        ri += 1

    data_end = ri - 1

    # Subtotal
    dat(ws2, ri, 1, f'Total {atc_label}', font=DARK_BOLD, fill=TOTAL_FILL)
    dat(ws2, ri, 2, '', fill=TOTAL_FILL)
    for ci in range(3, 3+len(yr_short)):
        c=ws2.cell(ri,ci,f'=SUM({get_column_letter(ci)}6:{get_column_letter(ci)}{data_end})')
        c.font=DARK_BOLD; c.alignment=RIGHT; c.fill=TOTAL_FILL; c.border=BORDER; c.number_format='#,##0'
    for ci in range(3+len(yr_short), 3+2*len(yr_short)):
        c=ws2.cell(ri,ci,f'=SUM({get_column_letter(ci)}6:{get_column_letter(ci)}{data_end})')
        c.font=DARK_BOLD; c.alignment=RIGHT; c.fill=TOTAL_FILL; c.border=BORDER; c.number_format='#,##0.000'
    if prev_yr:
        col_p = get_column_letter(3+len(yr_short)+yr_short.index(prev_yr))
        col_l = get_column_letter(3+len(yr_short)+yr_short.index(last_yr))
        crec_ci = 3+2*len(yr_short)
        c=ws2.cell(ri,crec_ci,f'=IF({col_p}{ri}=0,"",({col_l}{ri}-{col_p}{ri})/{col_p}{ri})')
        c.font=DARK_BOLD; c.alignment=RIGHT; c.fill=TOTAL_FILL; c.border=BORDER; c.number_format='0.00%'
    ri += 1

    # Total general
    c=ws2.cell(ri,1,'Total general'); c.font=WHITE_BOLD; c.fill=MED_BLUE; c.border=BORDER; c.alignment=LEFT
    ws2.cell(ri,2,'').fill=MED_BLUE; ws2.cell(ri,2).border=BORDER
    for ci in range(3, 3+len(yr_short)):
        c=ws2.cell(ri,ci,f'={get_column_letter(ci)}{ri-1}')
        c.font=WHITE_BOLD; c.fill=MED_BLUE; c.border=BORDER; c.alignment=RIGHT; c.number_format='#,##0'
    for ci in range(3+len(yr_short), 3+2*len(yr_short)):
        c=ws2.cell(ri,ci,f'={get_column_letter(ci)}{ri-1}')
        c.font=WHITE_BOLD; c.fill=MED_BLUE; c.border=BORDER; c.alignment=RIGHT; c.number_format='#,##0.000'
    if prev_yr:
        c=ws2.cell(ri,crec_ci,f'={get_column_letter(crec_ci)}{ri-1}')
        c.font=WHITE_BOLD; c.fill=MED_BLUE; c.border=BORDER; c.alignment=RIGHT; c.number_format='0.00%'

    ws2.column_dimensions['A'].width=36; ws2.column_dimensions['B'].width=55
    for ci in range(3, len(ct_hdrs)+1):
        ws2.column_dimensions[get_column_letter(ci)].width=16

    # ── HOJA 03_Mercado Molécula ──────────────────────────────────────────
    ws3 = wb.create_sheet('03_Mercado Molécula')
    filter_header(ws3)

    mol_hdrs = ['MERCADO','MOLECULA','MARCA','PRESENTACION'] + \
               [f'UNI {y}' for y in yr_short] + [f'US$ {y}' for y in yr_short]
    for ci, h in enumerate(mol_hdrs, 1):
        hdr(ws3, 5, ci, h)
    ws3.row_dimensions[5].height = 30

    grp = df_molecula.groupby(['Molecule Desc','Prod Desc','Pack Desc'])[uni_mat+usd_mat].sum().reset_index()
    grp = grp.sort_values(f'USD {last_yr}', ascending=False)

    ri = 6
    mol_first  = {}
    mol_ranges = {}

    for _, grow in grp.iterrows():
        mol  = grow['Molecule Desc']
        prod = grow['Prod Desc']
        pack = grow['Pack Desc']

        if mol not in mol_first:
            mol_first[mol]  = ri
            mol_ranges[mol] = []
        mol_ranges[mol].append(ri)

        fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL

        # MERCADO — repetido en cada fila
        dat(ws3, ri, 1, atc_label, fill=fill)
        # MOLECULA — repetida en cada fila (necesario para que SUMIFS funcione)
        dat(ws3, ri, 2, mol, fill=fill)

        dat(ws3, ri, 3, prod, fill=fill)
        dat(ws3, ri, 4, pack, fill=fill)

        # SUMIFS referencia B{ri} directamente — siempre tiene valor
        for ci, ul in enumerate(uni_letters, 5):
            c = ws3.cell(ri, ci, sumifs_mol('03_Mercado Molécula', ul, f'B{ri}', f'C{ri}', f'D{ri}'))
            c.font=DARK_NORM; c.alignment=RIGHT; c.fill=fill; c.border=BORDER; c.number_format='#,##0'
        for ci, dl in enumerate(usd_letters, 5+len(yr_short)):
            c = ws3.cell(ri, ci, sumifs_mol('03_Mercado Molécula', dl, f'B{ri}', f'C{ri}', f'D{ri}'))
            c.font=DARK_NORM; c.alignment=RIGHT; c.fill=fill; c.border=BORDER; c.number_format='#,##0.000'
        ri += 1

    # Subtotales por molécula
    grand_start = ri
    for mol, rows in mol_ranges.items():
        dat(ws3, ri, 1, '', fill=TOTAL_FILL)
        dat(ws3, ri, 2, f'Total {mol}', font=DARK_BOLD, fill=TOTAL_FILL)
        dat(ws3, ri, 3, '', fill=TOTAL_FILL)
        dat(ws3, ri, 4, '', fill=TOTAL_FILL)
        for ci in range(5, 5+len(yr_short)):
            refs = ','.join([f'{get_column_letter(ci)}{r}' for r in rows])
            c=ws3.cell(ri,ci,f'=SUM({refs})')
            c.font=DARK_BOLD; c.alignment=RIGHT; c.fill=TOTAL_FILL; c.border=BORDER; c.number_format='#,##0'
        for ci in range(5+len(yr_short), 5+2*len(yr_short)):
            refs = ','.join([f'{get_column_letter(ci)}{r}' for r in rows])
            c=ws3.cell(ri,ci,f'=SUM({refs})')
            c.font=DARK_BOLD; c.alignment=RIGHT; c.fill=TOTAL_FILL; c.border=BORDER; c.number_format='#,##0.000'
        ri += 1

    # Total ATC
    dat(ws3, ri, 1, f'Total {atc_label}', font=DARK_BOLD, fill=TOTAL_FILL)
    for cc in [2,3,4]: dat(ws3, ri, cc, '', fill=TOTAL_FILL)
    for ci in range(5, 5+len(yr_short)):
        c=ws3.cell(ri,ci,f'=SUM({get_column_letter(ci)}{grand_start}:{get_column_letter(ci)}{ri-1})')
        c.font=DARK_BOLD; c.alignment=RIGHT; c.fill=TOTAL_FILL; c.border=BORDER; c.number_format='#,##0'
    for ci in range(5+len(yr_short), 5+2*len(yr_short)):
        c=ws3.cell(ri,ci,f'=SUM({get_column_letter(ci)}{grand_start}:{get_column_letter(ci)}{ri-1})')
        c.font=DARK_BOLD; c.alignment=RIGHT; c.fill=TOTAL_FILL; c.border=BORDER; c.number_format='#,##0.000'
    ri += 1

    # Total general
    c=ws3.cell(ri,1,'Total general'); c.font=WHITE_BOLD; c.fill=MED_BLUE; c.border=BORDER; c.alignment=LEFT
    for cc in [2,3,4]: ws3.cell(ri,cc,'').fill=MED_BLUE; ws3.cell(ri,cc).border=BORDER
    for ci in range(5, 5+len(yr_short)):
        c=ws3.cell(ri,ci,f'={get_column_letter(ci)}{ri-1}')
        c.font=WHITE_BOLD; c.fill=MED_BLUE; c.border=BORDER; c.alignment=RIGHT; c.number_format='#,##0'
    for ci in range(5+len(yr_short), 5+2*len(yr_short)):
        c=ws3.cell(ri,ci,f'={get_column_letter(ci)}{ri-1}')
        c.font=WHITE_BOLD; c.fill=MED_BLUE; c.border=BORDER; c.alignment=RIGHT; c.number_format='#,##0.000'

    ws3.column_dimensions['A'].width=36; ws3.column_dimensions['B'].width=48
    ws3.column_dimensions['C'].width=22; ws3.column_dimensions['D'].width=34
    for ci in range(5, len(mol_hdrs)+1):
        ws3.column_dimensions[get_column_letter(ci)].width=15

    # ── HOJA FECHA LANZAMIENTO ────────────────────────────────────────────
    ws4 = wb.create_sheet('FECHA LANZAMIENTO')
    filter_header(ws4)

    fl_hdrs = ['MERCADO','MOLECULA','MARCA','FECHA LANZAMIENTO',
               f'UNI {last_yr}', f'US$ {last_yr}']
    for ci, h in enumerate(fl_hdrs, 1):
        hdr(ws4, 5, ci, h)
    ws4.row_dimensions[5].height = 30

    brands = df_molecula.groupby(['Molecule Desc','Prod Desc','Pack Launch Dt']).agg(
        **{f'UNI {last_yr}':(f'UNI {last_yr}','sum'),
           f'USD {last_yr}':(f'USD {last_yr}','sum')}
    ).reset_index().sort_values(f'USD {last_yr}', ascending=False)

    ri = 6
    for _, brow in brands.iterrows():
        mol  = brow['Molecule Desc']
        prod = brow['Prod Desc']
        fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL

        # MERCADO y MOLECULA repetidos en cada fila
        dat(ws4, ri, 1, atc_label, fill=fill)
        dat(ws4, ri, 2, mol, fill=fill)

        dat(ws4, ri, 3, prod, fill=fill)
        dat(ws4, ri, 4, fmt_launch(brow['Pack Launch Dt']), fill=fill)

        for ci, ul in [(5, uni_letters[-1]), (6, usd_letters[-1])]:
            fmt_n = '#,##0' if ci==5 else '#,##0.000'
            # En Fecha Lanzamiento col D es la fecha, no Pack Desc
            # Filtrar solo por Molecula + Producto (sin Pack)
            pais_ref  = "'FECHA LANZAMIENTO'!$B$1"
            forma_ref = "'FECHA LANZAMIENTO'!$B$2"
            data_val  = f"DATA!${ul}$2:DATA!${ul}${data_last_row}"
            formula = (
                f"=SUMIFS({data_val},"
                f"DATA!${L_MOL}$2:DATA!${L_MOL}${data_last_row},B{ri},"
                f"DATA!${L_PROD}$2:DATA!${L_PROD}${data_last_row},C{ri},"
                f"DATA!${L_PAIS}$2:DATA!${L_PAIS}${data_last_row},IF({pais_ref}=\"(Todos)\",\"*\",{pais_ref}),"
                f"DATA!${L_FORMA}$2:DATA!${L_FORMA}${data_last_row},IF({forma_ref}=\"(Todas)\",\"*\",{forma_ref}))"
            )
            c = ws4.cell(ri, ci, formula)
            c.font=DARK_NORM; c.alignment=RIGHT; c.fill=fill; c.border=BORDER; c.number_format=fmt_n
        ri += 1

    dat(ws4, ri, 1, f'Total {atc_label}', font=DARK_BOLD, fill=TOTAL_FILL)
    for cc in [2,3,4]: dat(ws4, ri, cc, '', fill=TOTAL_FILL)
    for ci in [5,6]:
        fmt_n = '#,##0' if ci==5 else '#,##0.000'
        c=ws4.cell(ri,ci,f'=SUM({get_column_letter(ci)}6:{get_column_letter(ci)}{ri-1})')
        c.font=DARK_BOLD; c.alignment=RIGHT; c.fill=TOTAL_FILL; c.border=BORDER; c.number_format=fmt_n
    ri += 1

    c=ws4.cell(ri,1,'Total general'); c.font=WHITE_BOLD; c.fill=MED_BLUE; c.border=BORDER; c.alignment=LEFT
    for cc in [2,3,4]: ws4.cell(ri,cc,'').fill=MED_BLUE; ws4.cell(ri,cc).border=BORDER
    for ci in [5,6]:
        fmt_n = '#,##0' if ci==5 else '#,##0.000'
        c=ws4.cell(ri,ci,f'={get_column_letter(ci)}{ri-1}')
        c.font=WHITE_BOLD; c.fill=MED_BLUE; c.border=BORDER; c.alignment=RIGHT; c.number_format=fmt_n

    ws4.column_dimensions['A'].width=36; ws4.column_dimensions['B'].width=45
    ws4.column_dimensions['C'].width=22; ws4.column_dimensions['D'].width=20
    ws4.column_dimensions['E'].width=16; ws4.column_dimensions['F'].width=18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── INTERFAZ STREAMLIT ────────────────────────────────────────────────────
with st.sidebar:
    st.header("📂 Sábana IQVIA")
    archivo = st.file_uploader("Sube el archivo mensual (.xlsx)", type=['xlsx'])
    if archivo:
        st.success("Archivo cargado ✓")

if not archivo:
    st.info("👈 Sube la sábana mensual de IQVIA en el panel izquierdo.")
    st.stop()

file_bytes = archivo.read()
df, unit_cols, usd_cols = cargar_sabana(file_bytes)
mat_periods = compute_mat_periods(unit_cols, usd_cols)

if not mat_periods:
    st.error("No se encontraron periodos MAT completos. Verifica que el archivo tenga al menos 12 meses.")
    st.stop()

# Mostrar tabla de periodos MAT detectados
st.success(f"✅ {len(df):,} registros cargados")

with st.expander("📅 Periodos MAT detectados", expanded=True):
    mat_data = []
    for lbl, ucols, dcols in mat_periods:
        inicio = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{4})', ucols[0])
        fin    = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{4})', ucols[-1])
        mat_data.append({
            'MAT': lbl,
            'Inicio': f"{inicio.group(1)} {inicio.group(2)}" if inicio else '?',
            'Fin':    f"{fin.group(1)} {fin.group(2)}"       if fin    else '?',
            'Meses':  len(ucols)
        })
    st.table(pd.DataFrame(mat_data))

st.divider()
st.subheader("🔍 Criterios de búsqueda")
st.caption("El Excel tendrá toda la clase terapéutica en DATA y dropdowns dinámicos de país y forma.")

col1, col2, col3 = st.columns(3)
with col1:
    sel_mol = st.multiselect("🧪 Molécula",
        options=sorted(df['Molecule Desc'].dropna().unique().tolist()),
        placeholder="Busca o selecciona moléculas...")
with col2:
    sel_prod = st.multiselect("💊 Producto / Marca",
        options=sorted(df['Prod Desc'].dropna().unique().tolist()),
        placeholder="Busca o selecciona productos...")
with col3:
    sel_atc = st.multiselect("📋 Clase terapéutica (ATC IV)",
        options=sorted(df['Atc IV'].dropna().unique().tolist()),
        placeholder="Busca o selecciona clase ATC...")

col4, col5 = st.columns(2)
paises_disp = sorted(df['Country Desc'].dropna().unique().tolist())
formas_disp = sorted(df['App1 Desc'].dropna().unique().tolist())
with col4:
    sel_pais = st.multiselect("🌎 País",
        options=paises_disp, default=paises_disp,
        placeholder="Selecciona países...")
with col5:
    sel_forma = st.multiselect("💉 Forma de administración",
        options=formas_disp, default=formas_disp,
        placeholder="Selecciona formas...")

st.caption("💡 El Excel incluye dropdowns en celdas amarillas para filtrar por país y forma directamente.")
st.divider()
generar = st.button("🚀 Generar Excel", type="primary", use_container_width=True)

if generar:
    if not any([sel_mol, sel_prod, sel_atc]):
        st.warning("⚠️ Selecciona al menos una molécula, producto o clase terapéutica.")
        st.stop()
    if not sel_pais:
        st.warning("⚠️ Selecciona al menos un país.")
        st.stop()

    # Filtro búsqueda
    mask_mol = pd.Series([True]*len(df), index=df.index)
    if sel_mol:  mask_mol &= df['Molecule Desc'].isin(sel_mol)
    if sel_prod: mask_mol &= df['Prod Desc'].isin(sel_prod)
    if sel_atc:  mask_mol &= df['Atc IV'].isin(sel_atc)
    mask_mol &= df['Country Desc'].isin(sel_pais)
    if len(sel_forma) < len(formas_disp):
        mask_mol &= df['App1 Desc'].isin(sel_forma)

    df_molecula = df[mask_mol].copy()
    if len(df_molecula) == 0:
        st.error("❌ No se encontraron registros.")
        st.stop()

    # ATC IV completo para DATA
    atc_iv = df_molecula['Atc IV'].mode()[0]
    mask_atc = df['Atc IV'] == atc_iv
    mask_atc &= df['Country Desc'].isin(sel_pais)
    if len(sel_forma) < len(formas_disp):
        mask_atc &= df['App1 Desc'].isin(sel_forma)
    df_atc = df[mask_atc].copy()

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Registros búsqueda",     f"{len(df_molecula):,}")
    c2.metric("Registros ATC completo", f"{len(df_atc):,}")
    c3.metric("Moléculas en ATC",       df_atc['Molecule Desc'].nunique())
    c4.metric("Países",                 df_molecula['Country Desc'].nunique())

    mol_buscada = sel_mol[0] if sel_mol else sel_prod[0] if sel_prod else atc_iv

    with st.spinner("Generando Excel..."):
        excel_buf = generar_excel(df_atc.copy(), df_molecula.copy(), mat_periods)

    nombre_base = re.sub(r'[^A-Za-z0-9]','_', mol_buscada)[:30].upper()
    last_mat    = mat_periods[-1][0].replace(' ','_')
    filename    = f"MERCADO_{nombre_base}_{last_mat}.xlsx"

    st.download_button(
        label="⬇️ Descargar Excel",
        data=excel_buf, file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
    st.success(f"✅ **{filename}** listo. Cambia las celdas amarillas en el Excel para filtrar.")
